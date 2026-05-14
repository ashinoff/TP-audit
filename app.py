"""
TP Audit — поиск абонентов, привязанных не к своей ТП.

Streamlit-приложение. Загружает Excel-выгрузку, анализирует распределение абонентов
по ТП и выдаёт ранжированный список подозрительных привязок. Данные не сохраняются.
"""
from __future__ import annotations

import io
import re
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from typing import Optional

import pandas as pd
import streamlit as st


# ═════════════════════════════════ КОНФИГУРАЦИЯ ═════════════════════════════════

st.set_page_config(
    page_title="TP Audit — аудит привязок к ТП",
    page_icon="◐",
    layout="wide",
    initial_sidebar_state="collapsed",
)


# ═════════════════════════════════ СХЕМА КОЛОНОК ═════════════════════════════════

@dataclass(frozen=True)
class FieldSpec:
    key: str
    label: str
    required: bool
    role: str
    exact: tuple[str, ...] = ()
    prefix: tuple[str, ...] = ()
    contains: tuple[str, ...] = ()
    avoid_contains: tuple[str, ...] = ()


SCHEMA: tuple[FieldSpec, ...] = (
    FieldSpec(
        key="tp", label="ТП", required=True,
        role="Идентификатор трансформаторной подстанции",
        exact=("тп", "тп номер", "номер тп", "наименование тп", "код тп", "tp"),
        prefix=("тп ",),
        contains=("трансформатор",),
        avoid_contains=("ту", "тус", "стек"),
    ),
    FieldSpec(
        key="pod", label="Подключение / ПС", required=False,
        role="Питающая вышестоящая подстанция",
        exact=("подключение", "пс", "питающая пс", "источник питания",
               "вышестоящая пс", "наименование пс", "питающая подстанция"),
        contains=("подключ", "питающ"),
    ),
    FieldSpec(
        key="feeder", label="Фидер", required=False,
        role="Линия 6/10 кВ от ПС до ТП",
        exact=("фидер", "присоединение"),
        contains=("фидер",),
    ),
    FieldSpec(
        key="np", label="Населённый пункт", required=False,
        role="Город, посёлок, село",
        exact=("населенный пункт", "населённый пункт", "город", "нп", "пункт"),
        contains=("населен", "населён"),
    ),
    FieldSpec(
        key="street", label="Улица", required=False,
        role="Название улицы или СНТ",
        exact=("улица", "ул"),
        contains=("улиц",),
    ),
    FieldSpec(
        key="house", label="Дом", required=False,
        role="Номер дома или участка",
        exact=("дом", "номер дома", "№ дома"),
    ),
    FieldSpec(
        key="ls", label="Лицевой счёт", required=False,
        role="Выводится в отчёт",
        exact=("лс", "лс / лс стек", "лс/лс стек", "лицевой счёт",
               "лицевой счет", "номер лс"),
        contains=("лицев",),
    ),
    FieldSpec(
        key="contract", label="Договор", required=False,
        role="Наименование договора / контрагент",
        exact=("наименование договора", "договор", "контрагент",
               "плательщик", "абонент"),
        contains=("договор", "контрагент"),
    ),
    FieldSpec(
        key="num", label="№ п/п", required=False,
        role="Порядковый номер строки",
        exact=("№ п/п", "n п/п", "n пп", "пп", "№", "n"),
    ),
    FieldSpec(
        key="tu_code", label="Код ТУ", required=False,
        role="Код точки учёта",
        exact=("код ту",),
    ),
    FieldSpec(
        key="tu_name", label="ТУ", required=False,
        role="Наименование точки учёта",
        exact=("ту", "наименование ту", "точка учёта", "точка учета"),
        avoid_contains=("стек", "код"),
    ),
)


# ═════════════════════════════════ НОРМАЛИЗАЦИЯ ═════════════════════════════════

_STREET_SUFFIX_RE = re.compile(
    r"\b(ул|улица|пер|переулок|пр|проспект|пр кт|пл|площадь|ш|шоссе|снт|с/т|сдт|"
    r"тер|тсн|днп|нп|кп|туп|тупик|б р|бульвар|кт|мкр|микрорайон)\b\.?",
    flags=re.IGNORECASE,
)
_PUNCT_RE = re.compile(r"[«»\"'.,\-()/]")
_SPACES_RE = re.compile(r"\s+")


def _norm_header(s: str) -> str:
    s = str(s)
    s = unicodedata.normalize("NFKC", s)
    s = s.lower().strip().replace("ё", "е")
    s = re.sub(r"[\u2010-\u2015_–—-]", " ", s)
    s = re.sub(r"[\.\"'`]", " ", s)
    s = _SPACES_RE.sub(" ", s).strip()
    return s


def detect_columns(df: pd.DataFrame) -> dict[str, Optional[str]]:
    headers = list(df.columns)
    normed = {h: _norm_header(h) for h in headers}

    candidates: list[tuple[int, str, str]] = []
    for field in SCHEMA:
        for col, n_col in normed.items():
            if any(av in n_col for av in field.avoid_contains):
                continue
            score = 0
            if any(_norm_header(a) == n_col for a in field.exact):
                score = 100
            elif any(n_col.startswith(_norm_header(a) + " ") or n_col == _norm_header(a)
                     for a in field.prefix):
                score = 80
            elif any(re.search(rf"\b{re.escape(_norm_header(a))}\b", n_col)
                     for a in (*field.exact, *field.contains)):
                score = 60
            elif any(_norm_header(a) in n_col for a in field.contains):
                score = 40
            if score > 0:
                candidates.append((score, field.key, col))

    candidates.sort(key=lambda x: (-x[0], x[1]))
    result: dict[str, Optional[str]] = {f.key: None for f in SCHEMA}
    used: set[str] = set()
    for _, field_key, col in candidates:
        if result[field_key] is None and col not in used:
            result[field_key] = col
            used.add(col)
    return result


def norm_street(s) -> Optional[str]:
    if pd.isna(s):
        return None
    s = str(s).lower().strip()
    s = _STREET_SUFFIX_RE.sub(" ", s)
    s = _PUNCT_RE.sub(" ", s)
    s = _SPACES_RE.sub(" ", s).strip()
    return s or None


def norm_house(h) -> Optional[str]:
    if pd.isna(h):
        return None
    s = str(h).lower().strip().replace(" ", "")
    return s or None


def parse_house_int(h) -> Optional[int]:
    """
    Извлекает целое число из обозначения дома.
    Примеры: "5" → 5, "5а" → 5, "10/2" → 10, "23:49:0402030:507" → None
    (кадастровые номера и слишком большие числа отсеиваются).
    """
    if pd.isna(h):
        return None
    s = str(h).strip()
    # Если в строке есть символ ":" — это, скорее всего, кадастровый номер
    if ":" in s:
        return None
    m = re.match(r"\s*(\d+)", s)
    if not m:
        return None
    try:
        n = int(m.group(1))
    except ValueError:
        return None
    # Разумный диапазон номера дома — отсеиваем мусор
    if 1 <= n <= 9999:
        return n
    return None


def norm_np(s) -> str:
    if pd.isna(s):
        return ""
    return str(s).lower().strip()


# ═════════════════════════════════ ПОДБОР КАНДИДАТОВ ═════════════════════════════════

def find_bracketing_tps(
    house_int: Optional[int],
    street: Optional[str],
    np_val: str,
    current_tp,
    *,
    houses_by_street_np_tp: dict,
    min_houses_on_other: int = 3,
    max_neighbor_distance: int = 15,
) -> list[tuple]:
    """
    Ищет ТП, у которых дом абонента «зажат» между их домами на той же улице.

    Возвращает [(tp, below, above, n_houses, total_distance), …] где
        below — ближайший дом ≤ house_int на этой ТП
        above — ближайший дом ≥ house_int
        n_houses — сколько всего домов этой улицы на той ТП
        total_distance — сумма расстояний до ближайших соседей (для ранжирования)
    """
    if house_int is None or street is None:
        return []
    out = []
    # Перебираем все (street, np, tp) с подходящей улицей в текущем нас. пункте
    for (s, n, tp_other), houses in houses_by_street_np_tp.items():
        if s != street or n != np_val or tp_other == current_tp:
            continue
        if len(houses) < min_houses_on_other:
            continue
        below = [h for h in houses if h <= house_int]
        above = [h for h in houses if h >= house_int]
        if not below or not above:
            continue
        closest_below = max(below)
        closest_above = min(above)
        # Точное совпадение считаем «слишком близко» — это уже сигнал 5
        if closest_below == house_int or closest_above == house_int:
            continue
        # Хотя бы один сосед должен быть в пределах max_neighbor_distance
        nearest = min(house_int - closest_below, closest_above - house_int)
        if nearest > max_neighbor_distance:
            continue
        total_distance = (house_int - closest_below) + (closest_above - house_int)
        out.append((tp_other, closest_below, closest_above, len(houses), total_distance))
    # Сортируем: чем меньше суммарное расстояние и чем больше домов — тем лучше
    out.sort(key=lambda x: (x[4], -x[3]))
    return out


def find_candidate_tps(
    row,
    current_tp,
    *,
    addr_tp: dict,
    street_np_tp: dict,
    tps_by_pod_feeder: dict,
    tps_by_feeder: dict,
    houses_by_street_np_tp: dict,
    max_candidates: int = 3,
) -> list[tuple[str, str]]:
    """
    Возвращает до `max_candidates` кортежей (tp_name, short_reason),
    отсортированных по приоритету:
        1. совпадение точного адреса (улица + дом) в этом же нас. пункте
        2. дом «зажат» соседями на другой ТП по этой улице
        3. совпадение улицы в том же нас. пункте на другой ТП
        4. совпадение ПС + фидера абонента с доминирующими у ТП
        5. совпадение только фидера (запасной вариант)
    """
    # (priority, secondary_sort, tp, reason)
    pool: list[tuple[int, int, str, str]] = []

    # P1 — точный адрес где-то ещё
    if pd.notna(row["_street"]) and pd.notna(row["_house"]):
        addr_c = addr_tp.get((row["_np"], row["_street"], row["_house"]), Counter())
        for tp, n in addr_c.most_common():
            if tp != current_tp and n >= 3:
                pool.append((1, -n, tp, f"адрес ×{n}"))

    # P2 — дом «зажат» соседями на другой ТП
    house_int = row.get("_house_int")
    if pd.notna(row["_street"]) and house_int is not None:
        brackets = find_bracketing_tps(
            house_int, row["_street"], row["_np"], current_tp,
            houses_by_street_np_tp=houses_by_street_np_tp,
        )
        for tp_other, below, above, _n, _dist in brackets:
            pool.append((2, _dist, tp_other, f"между домами №{int(below)} и №{int(above)}"))

    # P3 — улица в том же нас. пункте
    if pd.notna(row["_street"]):
        street_c = street_np_tp.get((row["_street"], row["_np"]), Counter())
        for tp, n in street_c.most_common():
            if tp != current_tp and n >= 5:
                pool.append((3, -n, tp, f"улица ×{n}"))

    # P4 — ПС + фидер совпадают
    if pd.notna(row["_pod"]) and pd.notna(row["_feeder"]):
        for tp in tps_by_pod_feeder.get((row["_pod"], row["_feeder"]), []):
            if tp != current_tp:
                pool.append((4, 0, tp, "ПС и фидер совпадают"))

    # P5 — только фидер
    if pd.notna(row["_feeder"]):
        for tp in tps_by_feeder.get(row["_feeder"], []):
            if tp != current_tp:
                pool.append((5, 0, tp, "фидер совпадает"))

    pool.sort()

    # Дедупликация — каждая ТП попадает один раз, с лучшим (наименьшим по приоритету) объяснением
    seen: set = set()
    out: list[tuple[str, str]] = []
    for _prio, _sec, tp, reason in pool:
        if tp in seen:
            continue
        seen.add(tp)
        out.append((tp, reason))
        if len(out) >= max_candidates:
            break
    return out


def _humanize_reason(reason_short: str) -> str:
    """Переводит компактный маркер кандидата в человеческий вид."""

    def _plural(n: int) -> str:
        """Корректное склонение слова «абонент» для русского числительного."""
        n = abs(n) % 100
        if 11 <= n <= 14:
            return "абонентов"
        n %= 10
        if n == 1:
            return "абонент"
        if 2 <= n <= 4:
            return "абонента"
        return "абонентов"

    if reason_short.startswith("адрес ×"):
        n = int(reason_short.replace("адрес ×", ""))
        return f"там уже {n} {_plural(n)} с тем же адресом"
    if reason_short.startswith("улица ×"):
        n = int(reason_short.replace("улица ×", ""))
        return f"там {n} {_plural(n)} с этой улицы"
    if reason_short.startswith("между домами"):
        return f"его номер дома {reason_short.replace('между домами', 'попадает между домами')} на этой ТП"
    if reason_short == "ПС и фидер совпадают":
        return "та же питающая ПС и тот же фидер — физически подходит"
    if reason_short == "фидер совпадает":
        return "тот же фидер"
    return reason_short


def format_candidates_humanized(candidates: list[tuple[str, str]]) -> str:
    """
    Преобразует список кандидатов в человеческое описание:
        - 1 кандидат:    «Скорее всего привязать к ТП-X — <причина>.»
        - 2-3 кандидата: «Скорее всего ТП-X — <причина>. Также возможно: ТП-Y (<причина>); ТП-Z (<причина>).»
    """
    if not candidates:
        return "Не удалось подобрать кандидатную ТП — проверьте вручную."

    first_tp, first_reason = candidates[0]
    head = f"Скорее всего привязать к {first_tp} — {_humanize_reason(first_reason)}"

    if len(candidates) == 1:
        return head + "."

    alts = []
    for tp, reason in candidates[1:]:
        alts.append(f"{tp} ({_humanize_reason(reason)})")
    return head + ". Менее уверенные варианты: " + "; ".join(alts) + "."


# ═════════════════════════════════ АНАЛИЗ ═════════════════════════════════

def analyze(
    df_raw: pd.DataFrame,
    cols: dict[str, Optional[str]],
    *,
    min_tp_size: int = 3,
    min_score: int = 30,
    weight_pod: int = 50,
    weight_feeder: int = 50,
    weight_np: int = 20,
    weight_street_orphan: int = 25,
    weight_addr_home: int = 30,
    weight_house_range: int = 25,
    np_dominance_threshold: float = 0.8,
) -> tuple[pd.DataFrame, dict]:
    if not cols.get("tp"):
        raise ValueError("Не выбрана колонка с ТП.")

    df = df_raw.copy()
    df["_tp"] = df[cols["tp"]]
    df["_feeder"] = df[cols["feeder"]] if cols.get("feeder") else pd.Series([None] * len(df), index=df.index)
    df["_pod"] = df[cols["pod"]] if cols.get("pod") else pd.Series([None] * len(df), index=df.index)
    df["_np_raw"] = df[cols["np"]] if cols.get("np") else pd.Series([None] * len(df), index=df.index)
    df["_street_raw"] = df[cols["street"]] if cols.get("street") else pd.Series([None] * len(df), index=df.index)
    df["_house_raw"] = df[cols["house"]] if cols.get("house") else pd.Series([None] * len(df), index=df.index)

    df["_np"] = df["_np_raw"].apply(norm_np)
    df["_street"] = df["_street_raw"].apply(norm_street)
    df["_house"] = df["_house_raw"].apply(norm_house)
    df["_house_int"] = df["_house_raw"].apply(parse_house_int)

    street_np_tp: dict[tuple, Counter] = defaultdict(Counter)
    addr_tp: dict[tuple, Counter] = defaultdict(Counter)
    houses_by_street_np_tp: dict[tuple, list[int]] = defaultdict(list)

    for _, row in df.iterrows():
        tp = row["_tp"]
        if pd.isna(tp):
            continue
        if pd.notna(row["_street"]):
            street_np_tp[(row["_street"], row["_np"])][tp] += 1
        if pd.notna(row["_street"]) and pd.notna(row["_house"]):
            addr_tp[(row["_np"], row["_street"], row["_house"])][tp] += 1
        if pd.notna(row["_street"]) and row["_house_int"] is not None:
            houses_by_street_np_tp[(row["_street"], row["_np"], tp)].append(row["_house_int"])

    # Сортируем списки домов для быстрой работы алгоритма «зажат»
    for key in houses_by_street_np_tp:
        houses_by_street_np_tp[key].sort()

    tp_modes: dict[str, dict] = {}
    for tp, sub in df.groupby("_tp"):
        if pd.isna(tp):
            continue
        feeder_top = sub["_feeder"].mode()
        pod_top = sub["_pod"].mode()
        np_top = sub["_np"].mode()
        np_top_val = np_top.iloc[0] if not np_top.empty else ""
        pod_top_val = pod_top.iloc[0] if not pod_top.empty else None
        feeder_top_val = feeder_top.iloc[0] if not feeder_top.empty else None
        # Берём первое оригинальное написание, соответствующее нормализованному np_top
        np_top_display = ""
        if np_top_val:
            matching = sub.loc[sub["_np"] == np_top_val, "_np_raw"]
            if not matching.empty:
                first_raw = matching.dropna()
                if not first_raw.empty:
                    np_top_display = str(first_raw.iloc[0])
        tp_modes[tp] = {
            "feeder_top": feeder_top_val,
            "pod_top": pod_top_val,
            "np_top": np_top_val,
            "np_top_display": np_top_display,
            "np_share": (sub["_np"] == np_top_val).mean() if np_top_val else 0.0,
            "pod_n": int((sub["_pod"] == pod_top_val).sum()) if pod_top_val else 0,
            "feeder_n": int((sub["_feeder"] == feeder_top_val).sum()) if feeder_top_val else 0,
            "streets": Counter(sub["_street"].dropna()),
            "size": len(sub),
        }

    # Индексы для подбора кандидатных ТП
    tps_by_feeder: dict[str, list[str]] = defaultdict(list)
    tps_by_pod_feeder: dict[tuple, list[str]] = defaultdict(list)
    for tp, m in tp_modes.items():
        if m["feeder_top"]:
            tps_by_feeder[m["feeder_top"]].append(tp)
            if m["pod_top"]:
                tps_by_pod_feeder[(m["pod_top"], m["feeder_top"])].append(tp)

    signal_counts = Counter()
    records = []

    for _, row in df.iterrows():
        tp = row["_tp"]
        if pd.isna(tp) or tp not in tp_modes:
            continue
        m = tp_modes[tp]
        if m["size"] < min_tp_size:
            continue

        score = 0
        reasons: list[str] = []

        if pd.notna(row["_pod"]) and m["pod_top"] and row["_pod"] != m["pod_top"]:
            score += weight_pod
            reasons.append(
                f"Питающая ПС не сходится: у абонента указана «{row['_pod']}», "
                f"а у {m['pod_n']} из {m['size']} абонентов этой ТП — «{m['pod_top']}». "
                f"Питающая подстанция у одной ТП должна быть одна — это ошибка в данных."
            )
            signal_counts["ПС"] += 1

        if pd.notna(row["_feeder"]) and m["feeder_top"] and row["_feeder"] != m["feeder_top"]:
            score += weight_feeder
            reasons.append(
                f"Фидер не сходится: у абонента «{row['_feeder']}», "
                f"а у {m['feeder_n']} из {m['size']} абонентов этой ТП — «{m['feeder_top']}». "
                f"Одна ТП питается одним фидером — физически невозможно."
            )
            signal_counts["Фидер"] += 1

        if (
            row["_np"]
            and m["np_top"]
            and row["_np"] != m["np_top"]
            and m["np_share"] >= np_dominance_threshold
        ):
            score += weight_np
            np_dom_display = m["np_top_display"] or m["np_top"]
            reasons.append(
                f"Адрес в населённом пункте «{row['_np_raw']}», но {m['np_share']:.0%} "
                f"абонентов этой ТП находятся в «{np_dom_display}». Похоже, нас. пункт "
                f"или ТП указаны с ошибкой — это разные территории."
            )
            signal_counts["Нас.пункт"] += 1

        if pd.notna(row["_street"]):
            count_here = m["streets"].get(row["_street"], 0)
            if count_here == 1:
                neighbours = street_np_tp.get((row["_street"], row["_np"]), Counter())
                others = Counter({k: v for k, v in neighbours.items() if k != tp})
                if others:
                    best_other_tp, best_other_n = others.most_common(1)[0]
                    dominant_count_here = (
                        m["streets"].most_common(1)[0][1] if m["streets"] else 0
                    )
                    if best_other_n >= 5 and dominant_count_here >= 5:
                        score += weight_street_orphan
                        reasons.append(
                            f"Из всех абонентов этой ТП только этот один — с улицы «{row['_street_raw']}» "
                            f"({row['_np_raw']}). При этом на ТП «{best_other_tp}» с той же улицы — "
                            f"{best_other_n} абонентов. Похоже, наш попал на эту ТП по ошибке."
                        )
                        signal_counts["Улица-одиночка"] += 1

        if pd.notna(row["_street"]) and pd.notna(row["_house"]):
            tps_for_addr = addr_tp.get((row["_np"], row["_street"], row["_house"]), Counter())
            total_at_addr = sum(tps_for_addr.values())
            here_at_addr = tps_for_addr.get(tp, 0)
            if total_at_addr >= 4 and (here_at_addr / total_at_addr) <= 0.25:
                best_tp, best_n = tps_for_addr.most_common(1)[0]
                if best_tp != tp and best_n >= 3:
                    score += weight_addr_home
                    others_count = total_at_addr - here_at_addr - best_n
                    tail = (
                        f", ещё {others_count} распределены иначе"
                        if others_count > 0 else ""
                    )
                    reasons.append(
                        f"По адресу «{row['_street_raw']} {row['_house_raw']}» в выгрузке "
                        f"{total_at_addr} абонентов: {best_n} сидят на ТП «{best_tp}»{tail}, "
                        f"и только этот один — на текущей ТП. Соседи по дому привязаны "
                        f"к другой подстанции."
                    )
                    signal_counts["Адрес"] += 1

        # Сигнал 6 — номер дома «зажат» между домами на другой ТП по этой улице
        if pd.notna(row["_street"]) and row["_house_int"] is not None:
            brackets = find_bracketing_tps(
                row["_house_int"], row["_street"], row["_np"], tp,
                houses_by_street_np_tp=houses_by_street_np_tp,
            )
            if brackets:
                score += weight_house_range
                best_tp_other, below, above, n_houses, _dist = brackets[0]
                reasons.append(
                    f"Дом №{int(row['_house_int'])} на улице «{row['_street_raw']}» "
                    f"({row['_np_raw']}) попадает в диапазон ТП «{best_tp_other}»: "
                    f"она держит здесь {n_houses} домов, ближайшие — №{int(below)} и №{int(above)}. "
                    f"По нумерации этот адрес должен быть на той же ТП."
                )
                signal_counts["Дом-диапазон"] += 1

        if score >= min_score:
            # Подбираем кандидатные ТП
            candidates = find_candidate_tps(
                row, tp,
                addr_tp=addr_tp,
                street_np_tp=street_np_tp,
                tps_by_pod_feeder=tps_by_pod_feeder,
                tps_by_feeder=tps_by_feeder,
                houses_by_street_np_tp=houses_by_street_np_tp,
            )
            candidates_str = format_candidates_humanized(candidates)

            # Формируем запись в нужном порядке:
            #   [инфо об абоненте] · Текущая ТП · Предлагаемая ТП · Причины · Балл
            rec: dict = {}

            # 1. Идентификационные поля абонента (только если есть в исходнике)
            for key in ("num", "ls", "contract", "tu_code", "tu_name"):
                col = cols.get(key)
                if col:
                    rec[col] = row[col]

            # 2. Адрес
            rec["Нас. пункт"] = row["_np_raw"]
            rec["Улица"] = row["_street_raw"]
            rec["Дом"] = row["_house_raw"]

            # 3. Текущие энергетические параметры абонента
            rec["Подключение"] = row["_pod"]
            rec["Фидер"] = row["_feeder"]

            # 4. Текущая и предлагаемая ТП
            rec["Текущая ТП"] = tp
            rec["Предлагаемая ТП"] = candidates_str

            # 5. Обоснование и итоговый балл
            rec["Причины"] = "; ".join(reasons)
            rec["Балл"] = score

            records.append(rec)

    result = pd.DataFrame(records)
    if not result.empty:
        result = result.sort_values(["Балл", "Текущая ТП"], ascending=[False, True]).reset_index(drop=True)

    summary = {
        "rows_total": len(df),
        "tps_total": df["_tp"].nunique(dropna=True),
        "tps_analyzed": sum(1 for m in tp_modes.values() if m["size"] >= min_tp_size),
        "suspicious_total": len(result),
        "signal_counts": dict(signal_counts),
    }
    return result, summary


def per_tp_summary(result: pd.DataFrame) -> pd.DataFrame:
    if result.empty:
        return pd.DataFrame()
    return (
        result.groupby("Текущая ТП")
        .agg(
            Подозрительных=("Балл", "size"),
            Сумма_баллов=("Балл", "sum"),
            Макс_балл=("Балл", "max"),
        )
        .sort_values(["Макс_балл", "Сумма_баллов"], ascending=False)
        .reset_index()
    )


def make_xlsx_bytes(result: pd.DataFrame, summary: dict, tp_agg: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        (result if not result.empty else pd.DataFrame({"Результат": ["Ничего не найдено"]})).to_excel(
            writer, sheet_name="Подозрительные", index=False
        )
        (tp_agg if not tp_agg.empty else pd.DataFrame({"Результат": ["Ничего не найдено"]})).to_excel(
            writer, sheet_name="По ТП", index=False
        )
        sig = summary.get("signal_counts", {})
        stats_rows = [
            ("Всего строк во входе", summary.get("rows_total", 0)),
            ("Всего ТП", summary.get("tps_total", 0)),
            ("ТП в анализе", summary.get("tps_analyzed", 0)),
            ("Подозрительных записей", summary.get("suspicious_total", 0)),
            ("— по сигналу «ПС»", sig.get("ПС", 0)),
            ("— по сигналу «Фидер»", sig.get("Фидер", 0)),
            ("— по сигналу «Нас.пункт»", sig.get("Нас.пункт", 0)),
            ("— по сигналу «Улица-одиночка»", sig.get("Улица-одиночка", 0)),
            ("— по сигналу «Адрес»", sig.get("Адрес", 0)),
            ("— по сигналу «Дом-диапазон»", sig.get("Дом-диапазон", 0)),
        ]
        pd.DataFrame(stats_rows, columns=["Показатель", "Значение"]).to_excel(
            writer, sheet_name="Статистика", index=False
        )

        from openpyxl.utils import get_column_letter
        for sheet_name in writer.book.sheetnames:
            ws = writer.book[sheet_name]
            for col_idx, col in enumerate(ws.columns, start=1):
                max_len = 0
                for cell in col:
                    v = cell.value
                    if v is not None:
                        max_len = max(max_len, min(len(str(v)), 80))
                ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2
    return buf.getvalue()


# ═════════════════════════════════ СТИЛИ ═════════════════════════════════

CUSTOM_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Fraunces:opsz,wght@9..144,300..600&family=Instrument+Sans:wght@400..600&family=JetBrains+Mono:wght@400;500&display=swap');

:root {
    --bg: #FAF7F0;
    --ink: #1A1614;
    --ink-soft: #6B6660;
    --ink-mute: #A39E97;
    --rule: #E0D9C7;
    --accent: #B8540C;
}

header[data-testid="stHeader"] { display: none; }
#MainMenu, footer { visibility: hidden; }
.stDeployButton { display: none; }

.stApp {
    background: #FAF7F0;
    color: var(--ink);
    font-family: 'Instrument Sans', system-ui, sans-serif;
}

.main .block-container {
    padding-top: 3rem;
    padding-bottom: 5rem;
    max-width: 900px;
}

h1, h2, h3 {
    font-family: 'Fraunces', serif !important;
    font-weight: 400;
    letter-spacing: -0.02em;
    color: var(--ink);
}

/* Хиро — компактный */
.pa-title {
    font-family: 'Fraunces', serif;
    font-size: 2.6rem;
    font-weight: 400;
    letter-spacing: -0.03em;
    line-height: 1;
    color: var(--ink);
    margin: 0;
}
.pa-title em {
    font-style: italic;
    color: var(--accent);
}
.pa-sub {
    font-size: 1rem;
    color: var(--ink-soft);
    margin: 0.5rem 0 0 0;
    max-width: 40em;
    line-height: 1.5;
}
.pa-divider {
    border: none;
    border-top: 1px solid var(--rule);
    margin: 2rem 0 1.5rem 0;
}

/* Метрики */
[data-testid="stMetric"] {
    background: transparent;
    border: none;
    padding: 0;
}
[data-testid="stMetricLabel"] {
    font-family: 'JetBrains Mono', monospace !important;
    font-size: 0.7rem !important;
    text-transform: uppercase;
    letter-spacing: 0.1em;
    color: var(--ink-mute) !important;
}
[data-testid="stMetricValue"] {
    font-family: 'Fraunces', serif !important;
    font-size: 2rem !important;
    color: var(--ink) !important;
    font-weight: 400 !important;
    letter-spacing: -0.02em;
    line-height: 1.2 !important;
}

/* Кнопки */
.stButton button, .stDownloadButton button {
    font-family: 'Instrument Sans', sans-serif !important;
    font-weight: 500 !important;
    border-radius: 2px !important;
    border: 1px solid var(--ink) !important;
    background: var(--ink) !important;
    color: var(--bg) !important;
    padding: 0.6rem 1.3rem !important;
    transition: all 0.15s ease;
}
.stButton button:hover, .stDownloadButton button:hover {
    background: var(--accent) !important;
    border-color: var(--accent) !important;
}

/* Аплоадер */
[data-testid="stFileUploader"] section {
    background: #FFFCF5;
    border: 1px dashed var(--ink-soft) !important;
    border-radius: 2px !important;
    padding: 1.5rem !important;
}
[data-testid="stFileUploader"] section:hover {
    border-color: var(--accent) !important;
}

/* Табы */
.stTabs [data-baseweb="tab-list"] {
    gap: 1.5rem;
    border-bottom: 1px solid var(--rule);
    background: transparent;
}
.stTabs [data-baseweb="tab"] {
    border-radius: 0 !important;
    background: transparent !important;
    color: var(--ink-soft) !important;
    font-family: 'Instrument Sans', sans-serif !important;
    font-size: 0.92rem !important;
    padding: 0.5rem 0 !important;
    border-bottom: 2px solid transparent !important;
    margin-bottom: -1px !important;
}
.stTabs [aria-selected="true"] {
    color: var(--ink) !important;
    border-bottom-color: var(--accent) !important;
}

[data-testid="stDataFrame"] {
    border: 1px solid var(--rule);
    border-radius: 2px;
}

.stAlert {
    border-radius: 2px !important;
    border-left: 3px solid var(--accent);
    background: #FFFCF5 !important;
}

.stSlider [data-baseweb="slider"] [role="slider"] {
    background: var(--accent) !important;
    border-color: var(--ink) !important;
}

/* Сайдбар */
section[data-testid="stSidebar"] {
    background: #F2EDE0;
    border-right: 1px solid var(--rule);
}

/* Раскрывающиеся блоки — поспокойнее */
.streamlit-expanderHeader {
    font-family: 'Instrument Sans', sans-serif !important;
    font-size: 0.92rem !important;
    color: var(--ink-soft) !important;
}

/* Уменьшаем шум от лейблов виджетов */
.stTextInput label, .stSelectbox label, .stSlider label, .stNumberInput label {
    font-size: 0.85rem !important;
    color: var(--ink-soft) !important;
}

/* Маленькая капитель для подписей */
.pa-tag {
    font-family: 'JetBrains Mono', monospace;
    font-size: 0.7rem;
    color: var(--ink-mute);
    letter-spacing: 0.1em;
    text-transform: uppercase;
}
</style>
"""


# ═════════════════════════════════ UI ═════════════════════════════════

def render_hero() -> None:
    st.markdown(
        """
        <div>
            <h1 class="pa-title">TP <em>Audit</em></h1>
            <p class="pa-sub">
                Поиск абонентов, которые с большой вероятностью привязаны не к своей ТП.
                Загрузите выгрузку — получите ранжированный список с обоснованием.
                Данные не сохраняются.
            </p>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_landing_help() -> None:
    """Минимальная справка перед загрузкой — всё в раскрывающихся блоках."""
    with st.expander("Какие колонки нужны во входном файле"):
        st.markdown(
            "Колонки распознаются автоматически по их заголовкам — порядок и точные "
            "названия не важны. **Обязательна только колонка с ТП.** "
            "Чем больше остальных полей передано — тем точнее анализ."
        )
        rows = []
        for f in SCHEMA:
            req = "обязательно" if f.required else "желательно"
            rows.append({"Колонка": f.label, "Назначение": f.role, "Статус": req})
        st.dataframe(pd.DataFrame(rows), hide_index=True, use_container_width=True)

    with st.expander("Как считаются баллы"):
        st.markdown(
            "Для каждого абонента складываются баллы по пяти независимым сигналам. "
            "Чем больше сработало — тем выше уверенность в ошибке."
        )
        method = pd.DataFrame(
            [
                {"Сигнал": "ПС", "Балл": 50,
                 "Логика": "Питающая ПС ≠ доминирующей на ТП. Физически невозможно."},
                {"Сигнал": "Фидер", "Балл": 50,
                 "Логика": "Фидер ≠ доминирующему на ТП. Физически невозможно."},
                {"Сигнал": "Нас. пункт", "Балл": 20,
                 "Логика": "Нас. пункт ≠ доминирующему, если тот охватывает ≥ 80 %."},
                {"Сигнал": "Улица", "Балл": 25,
                 "Логика": "Улица единственная на ТП, но имеет «домашнюю» ТП в этом же нас. пункте."},
                {"Сигнал": "Адрес", "Балл": 30,
                 "Логика": "Связка «улица + дом» в основном относится к другой ТП."},
                {"Сигнал": "Дом", "Балл": 25,
                 "Логика": "Номер дома попадает в диапазон домов другой ТП на этой улице."},
            ]
        )
        st.dataframe(method, hide_index=True, use_container_width=True)
        st.markdown(
            "**≥ 100** — почти наверняка ошибка &nbsp;·&nbsp; "
            "**50 – 99** — высокая вероятность &nbsp;·&nbsp; "
            "**30 – 49** — кандидат на проверку"
        )


def main() -> None:
    st.markdown(CUSTOM_CSS, unsafe_allow_html=True)
    render_hero()

    # ── Сайдбар ──────────────────────────────────────────────────────────────
    with st.sidebar:
        st.markdown('<div class="pa-tag">Параметры</div>', unsafe_allow_html=True)
        min_score = st.slider("Минимальный балл для отчёта", 20, 150, 30, step=5)
        min_tp_size = st.slider(
            "Минимум абонентов на ТП", 2, 20, 3,
            help="ТП с меньшим числом абонентов пропускаются.",
        )

        with st.expander("Дополнительно"):
            np_dom = st.slider(
                "Доминирование нас. пункта", 0.5, 1.0, 0.8, step=0.05,
                help="Доля абонентов в доминирующем нас. пункте, чтобы сигнал сработал.",
            )
            st.markdown('<div class="pa-tag" style="margin-top:0.8rem;">Веса сигналов</div>', unsafe_allow_html=True)
            w_pod = st.number_input("ПС", 0, 200, 50, step=5)
            w_feeder = st.number_input("Фидер", 0, 200, 50, step=5)
            w_np = st.number_input("Нас. пункт", 0, 200, 20, step=5)
            w_street = st.number_input("Улица-одиночка", 0, 200, 25, step=5)
            w_addr = st.number_input("Адрес у другой ТП", 0, 200, 30, step=5)
            w_house_range = st.number_input("Дом в диапазоне", 0, 200, 25, step=5)

    st.markdown('<hr class="pa-divider">', unsafe_allow_html=True)

    # ── Загрузка ─────────────────────────────────────────────────────────────
    uploaded = st.file_uploader(
        "Excel-файл (.xlsx)",
        type=["xlsx", "xlsm"],
        accept_multiple_files=False,
    )

    if uploaded is None:
        render_landing_help()
        return

    try:
        with st.spinner("Читаю файл…"):
            xls = pd.ExcelFile(uploaded)
            sheet_name = xls.sheet_names[0]
            if len(xls.sheet_names) > 1:
                sheet_name = st.selectbox("Лист", xls.sheet_names, index=0)
            df_raw = pd.read_excel(xls, sheet_name=sheet_name)
    except Exception as exc:
        st.error(f"Не удалось прочитать файл: {exc}")
        return

    auto_cols = detect_columns(df_raw)
    required_missing = [f.label for f in SCHEMA if f.required and not auto_cols.get(f.key)]

    # Если автодетект не нашёл ТП — обязательно показываем сопоставление
    expand_mapping = bool(required_missing)

    with st.expander("Сопоставление колонок (если автоопределение ошиблось)", expanded=expand_mapping):
        st.caption("Колонка ТП обязательна. Остальные — по желанию.")
        options = [None] + list(df_raw.columns)
        manual_cols = {}
        c1, c2, c3 = st.columns(3)
        for i, field in enumerate(SCHEMA):
            target = (c1, c2, c3)[i % 3]
            with target:
                default = auto_cols.get(field.key)
                idx = options.index(default) if default in options else 0
                label = f"{field.label}{' *' if field.required else ''}"
                manual_cols[field.key] = st.selectbox(
                    label, options, index=idx, key=f"col_{field.key}",
                )
        cols = manual_cols

    if not cols.get("tp"):
        st.error("Не указана колонка с ТП. Раскройте блок сопоставления и выберите её вручную.")
        return

    # ── Анализ ───────────────────────────────────────────────────────────────
    with st.spinner("Анализирую…"):
        try:
            result, summary = analyze(
                df_raw, cols,
                min_tp_size=min_tp_size,
                min_score=min_score,
                weight_pod=w_pod,
                weight_feeder=w_feeder,
                weight_np=w_np,
                weight_street_orphan=w_street,
                weight_addr_home=w_addr,
                weight_house_range=w_house_range,
                np_dominance_threshold=np_dom,
            )
        except Exception as exc:
            st.error(f"Ошибка анализа: {exc}")
            return

    st.markdown('<hr class="pa-divider">', unsafe_allow_html=True)

    # Метрики — 4 штуки, без второго ряда
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Строк", f"{summary['rows_total']:,}".replace(",", " "))
    m2.metric("ТП всего", f"{summary['tps_total']:,}".replace(",", " "))
    m3.metric("В анализе", f"{summary['tps_analyzed']:,}".replace(",", " "))
    m4.metric("Подозрительных", f"{summary['suspicious_total']:,}".replace(",", " "))

    if result.empty:
        st.warning("Ничего не найдено при текущих настройках. Снизьте порог балла в боковой панели.")
        return

    tp_agg = per_tp_summary(result)

    st.markdown('<div style="height:1.5rem;"></div>', unsafe_allow_html=True)

    tab1, tab2 = st.tabs(["Подозрительные", "По ТП"])

    with tab1:
        f1, f2 = st.columns([2, 1])
        with f1:
            tp_filter = st.text_input("Фильтр по ТП", "", placeholder="например, ТП-А216")
        with f2:
            min_show = st.slider("Балл от", min_score, 200, min_score, step=5, key="min_show")
        view = result.copy()
        if tp_filter:
            view = view[view["Текущая ТП"].astype(str).str.contains(tp_filter, case=False, na=False)]
        view = view[view["Балл"] >= min_show]
        st.dataframe(view, hide_index=True, use_container_width=True, height=500)

    with tab2:
        st.dataframe(tp_agg, hide_index=True, use_container_width=True, height=500)

    st.markdown('<div style="height:1rem;"></div>', unsafe_allow_html=True)

    xlsx_bytes = make_xlsx_bytes(result, summary, tp_agg)
    st.download_button(
        "Скачать отчёт",
        data=xlsx_bytes,
        file_name=f"tp-audit_{pd.Timestamp.today().strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )


if __name__ == "__main__":
    main()
